# -*- coding: utf-8 -*-#
# @Time :2019/4/1511:20
# @Author :xuqiao  
# @Email :1462942304@qq.com
# @File :read_write_excel.py
from openpyxl import load_workbook
from webservice_api.conmon import cantins
class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.api = None
        self.url = None
        self.data = None
        self.exp = None
        self.actual = None
        self.result = None
        self.sql = None
class ExcelTest:
    def __init__(self,file_name,sheet_name):
        try:
            self.file_name = file_name
            self.wb = load_workbook(file_name)
            self.sheet_name = self.wb[sheet_name]

        except Exception as e:
            print("输入的文件或sheet_name不正确")
    def read_excel(self):
        cases=[]
        for r in range(2,self.sheet_name.max_row+1):
            case = Case()
            case.case_id = self.sheet_name.cell(r,1).value
            case.title = self.sheet_name.cell(r,2).value
            case.api = self.sheet_name.cell(r,3).value
            case.url= self.sheet_name.cell(r,4).value
            case.data = self.sheet_name.cell(r,5).value
            case.exp = self.sheet_name.cell(r,6).value
            case.sql = self.sheet_name.cell(r,9).value
            cases.append(case)
        self.wb.close()
        return cases
    def write_excel(self,row,actual,result):
        self.sheet_name.cell(row,7).value = actual
        self.sheet_name.cell(row,8).value = result
        self.wb.save(self.file_name)
        self.wb.close()
if __name__ == '__main__':
    pass
    # res = ExcelTest(cantins.data_path,"register").read_excel()
    # ht_request = HttpRquests()
    # for case in res:
    #     print(case.__dict__)
        # rep = ht_request.http_res(case.method,case.url,case.data)
        # print(rep.text)
        # rep = rep.text
        # print(rep)
        # if case.exp == rep:
        #     ExcelTest(cantins.data_path, "register").write_excel(case.case_id+1,rep,"PASS")
        # else:
        #     ExcelTest(cantins.data_path, "register").write_excel(case.case_id + 1, rep, "FAIL")





